#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Sat Sep  9 17:13:13 2023

@author: michal
"""


import pandas as pd
import random as rn
from classes import cutting_df, TimeFormatModifier
from classes import file_n
import datetime
from openpyxl import load_workbook
from classes import def_mnth, pt, pt_to_save


current_month = datetime.date.today().month-1
current_year = datetime.date.today().year
if int(current_month) < 10:
    current_month = f'0{current_month}'
    
    
if int(current_month) == 1:
    current_month = '12'
    current_year -= 1


# Function to check if all lists in the list of lists are empty
def all_lists_empty(lists):
    return all(not lst for lst in lists)


path_input = input('Path to file with name: ')


sheet_general = str('Zestawienie')
worker_sheet = 'Prac 1'
# Define a search function
def search_string(s, search):
    return search in str(s).lower()

# Search for the string 'al' in all columns
sub = 'C00'
sub1 = "Wew"

user_input = input("The same NAME of WORKER as in Excel file: ")

with open(pt+'user_input.txt', 'w') as file:
    file.write(user_input)
    
with open(pt+'path_input.txt', 'w') as file:
    file.write(path_input)

if def_mnth == 1:
    month_year = input("Miesiac i rok: ")
    with open(pt+'month_year.txt', 'w') as file:
        file.write(month_year)
        
    current_month, current_year = month_year.split(" ")
    if int(current_month) < 10:
        current_month = f'0{current_month}'


new_dict = {}
file_name = f'{path_input}'
df = pd.read_excel(io=file_name, sheet_name=sheet_general)
copy_df = df.copy()

day_of_the_week = (df.iloc[6:7])


non_zero_columns = []
aaa = df.iloc[-1:]
# Iterate through the columns of the DataFrame
aaa = aaa.dropna(axis=1)


for column_name in aaa.columns:
    # Use the .loc indexer to select the column and check if any value is different from 0
    if (isinstance(aaa[column_name].iloc[0], int) and aaa[column_name].iloc[0] !=0) :
        non_zero_columns.append(column_name)
    elif (isinstance(aaa[column_name].iloc[0], float) and aaa[column_name].iloc[0] !=0):
        non_zero_columns.append(column_name)

get_cut = int(df.columns.get_loc(non_zero_columns[-1]) - len(df.columns))

weekends = []

for col in day_of_the_week:
    if day_of_the_week[col].values[0] == 6:
        
        weekends.append(col)
        
    elif day_of_the_week[col].values[0] == 7:
        weekends.append(col)


mm = cutting_df(0, 6, get_cut+1, day_of_the_week)
mm.dropin(mm.start_columns, mm.end_columns, mm.last_col, mm.df_name)
day_of_the_week = day_of_the_week.astype('int64')

df_worker = pd.read_excel(io=file_name, sheet_name=worker_sheet)
df_worker = df_worker[df_worker.ne(-1).all(axis=1)]

roboczo_godziny = pd.read_excel(io=file_n)
roboczo_godziny.set_index('Unnamed: 0', inplace = True)

f_name, s_name = user_input.split(' ')
user_input_reverse = s_name+' '+f_name
hours_project = list(roboczo_godziny[user_input_reverse]
                     .dropna()
                     )

names_of_projects = list(roboczo_godziny[user_input_reverse]
                     .dropna()
                     .index)
names_of_projects = ['praca na etacie' if isinstance(x, float) else x for x in names_of_projects]


if (sum(hours_project)-hours_project[-1]*2) == 0:
    names_of_projects.pop()
    hours_project.pop()




df["Index"] = df[df.columns[1]].eq(user_input) #give boolean series attached to orginal df
index = df["Index"]
name_index = index[index].index.values[0] #give index of true values




df_shorten = cutting_df(0, 6, get_cut, df_name=df)
df_shorten.dropin(df_shorten.start_columns, 
                  df_shorten.end_columns, 
                  df_shorten.last_col, 
                  df_shorten.df_name
                  )


day_month = df.iloc[5:6].astype('int64', copy=True)


weekende_date = day_month[weekends].values[0]



#hours of work in working days
hour_per_day = df.iloc[name_index:name_index+1]

cut_zest_per_worker = copy_df.iloc[name_index:name_index+7]

#Find individual number of worker


nr_str = '  Nr ewid.:'
mask = cut_zest_per_worker.eq(nr_str)
row_index = mask.any(axis=1).idxmax()
col_index = mask.any().sum()+1

nr_ewid = cut_zest_per_worker[cut_zest_per_worker.columns[col_index]].loc[row_index]


hour_per_day = (hour_per_day.dropna(axis=1))
columns_of_project_days = hour_per_day.columns
hour_per_day = (hour_per_day.astype('float', copy=True)
                )


for _ in range(len(hours_project)):
    new_dict[names_of_projects[_]] = hours_project[_]
    
new_dict = { k.replace('-', '_'): v for k, v in new_dict.items() }



keys = []
values = []
time_interval = 0.5

for i in new_dict:
    keys.append(i)
    values.append(new_dict[i])
    

    
list_of_projects = []
to_ad = 0
project_list_with_hours = []

for k, v in new_dict.items():
    
    
    k = k.replace('-', '_')
    globals()[k] = []

    interval_count = int(v/time_interval)
    [globals()[k].append(time_interval) for _ in range(interval_count)]
    
    project_list_with_hours.append(globals()[k])


full_list = []
full_list = []
only_hours = []
hours_with_index = {}


aa = 0
while not all_lists_empty(project_list_with_hours):
    name_of_project = rn.choice(keys)
    list_per_project = globals()[name_of_project]
    
    
    len_of_day = hour_per_day.values[0][_]*2
    


    try:
        if len(list_per_project) >= len_of_day:
            rand_numb = rn.randint(1,len_of_day)
        elif len(list_per_project) == 2:
            rand_numb=2
        elif len(list_per_project) == 3:
            rand_numb=3
        elif len(list_per_project) == 4:
            rand_numb=4
        else:
            rand_numb = rn.randint(1, int(len_of_day))
        
    
        choice_test = sum(rn.sample(list_per_project, rand_numb))
        aa+=choice_test

        
        full_list.append([name_of_project, choice_test])
        
        only_hours.append(choice_test)
        
        hours_with_index[name_of_project] = choice_test
    
            
        if len(list_per_project) == 0:
            keys.remove(name_of_project)

        del list_per_project[0:rand_numb]
        
        
        
    except ValueError:
        pass
    


df_list_projects = []
df_hours_projects = []
names_of_days = []


working_days = [day_month[column].values[0] for column in columns_of_project_days]

for k in working_days:
    string_int = str(k)
    string_all = string_int
    names_of_days.append(string_all)
    
dict_day_hour = {}
    
for name in names_of_days:
     dict_day_hour[name] = []
     
     
full_list_copy = full_list.copy()



input_dict = dict_day_hour.copy()

hours_list = full_list_copy
hour_per_day_values = list(hour_per_day.values[0])
remaining_hours = {}
for i in range(len(names_of_days)):
    remaining_hours[int(names_of_days[i])] = hour_per_day_values[i]
    

 

# Function to calculate the total hours in a list of lists
def calculate_total_hours(lst):
    total = 0
    for entry in lst:
        total += entry[1]
    return total



# Sort hours_list in descending order of hours_needed
hours_list.sort(key=lambda x: -x[1])


for (day, hours), (k, v) in zip(input_dict.items(), remaining_hours.items()):
        remaining_hours = v - calculate_total_hours(hours)
    
        while remaining_hours > 0:
            for i, entry in enumerate(hours_list):
                task, hours_needed = entry
                if hours_needed <= remaining_hours:
                    input_dict[day].append(entry)
                    remaining_hours -= hours_needed
                    hours_list.pop(i)

                if remaining_hours == 0:
                    break

data = input_dict

# Initialize a dictionary to store the aggregated data
aggregated_data = {}

# Iterate over the original data dictionary
for day, day_data in data.items():
    # Initialize a temporary dictionary to aggregate data for the current day
    temp_data = {}
    
    # Iterate over the entries for the current day
    for entry in day_data:
        project, value = entry
        
        # Check if the project exists in the temporary dictionary
        if project in temp_data:
            # If it exists, add the value to the existing value
            temp_data[project] += value
        else:
            
            # If it doesn't exist, create a new entry
            temp_data[project] = value
    
    # Convert the temporary dictionary to a list of [project, value] pairs
    aggregated_day_data = [[project, value] for project, value in temp_data.items()]
    
    # Update the aggregated_data dictionary with the aggregated data for the current day
    aggregated_data[day] = aggregated_day_data



# Initialize an empty list to store the rows
rows = []

# Iterate over the dictionary items and flatten the data
for day, day_data in aggregated_data.items():
    for project, value in day_data:
        rows.append([day, project, value])

# Create a DataFrame from the list of rows
df = pd.DataFrame(rows, columns=['day', 'project', 'value'])
df_val = df.copy()


# Your original dictionary
data = aggregated_data
import pandas as pd


dfs = []


cut_zest_per_worker = copy_df.iloc[name_index:name_index+7]

keys_for_worker = list(cut_zest_per_worker['Zestawienie zbiorcze'].values)
all_keys = list(df_worker[df_worker.columns[40]].iloc[9:36].values)


keys_rows = {}
for k in keys_for_worker:

    row_index, col_name = (cut_zest_per_worker == k).values.nonzero()
    
    # Check if the value was found
    if len(row_index) >= 0:
        # Convert the row and column indices to integers (assuming there's only one occurrence)
        row_index = int(row_index[0])
        col_name = cut_zest_per_worker.columns[col_name[0]]
        
        keys_rows[k] = row_index
        
    
    else:
        break

temp = cutting_df(0, 6, -4, cut_zest_per_worker)
temp.dropin(0, 6, -4, cut_zest_per_worker)

dict_other = {}
df_keys = pd.DataFrame()
dfs_to_concatenate = []
index_column = []

for k, v in keys_rows.items():
    if k == 'C1':
        pass
    else:
        klm = cut_zest_per_worker.iloc[v:v+1]
        klm = klm.dropna(axis=1)
        if not klm.empty:
            dfs_to_concatenate.append(klm)
            index_column.append(k)

# Concatenate all non-empty DataFrames in the list
if dfs_to_concatenate:
    df_keys = pd.concat(dfs_to_concatenate, ignore_index=True)
    
df_keys.index = index_column
        
    
# second_change_ind = []
second_change_indicator = 'II'
df_keys_copy = df_keys.copy()


if second_change_indicator in df_keys.values:
    
    df_keys_copy = df_keys.iloc[0:1].dropna(axis=1).astype(str)
    
    mask = df_keys_copy.eq(second_change_indicator)
    columns_second = df_keys_copy.columns[mask.any()].tolist()
    
    day_of_second_change = list(day_month[columns_second].values[0])
else:
    day_of_second_change = []
    pass



data_cop = data.copy()
# data = data_cop

dict_another = []
def is_int_or_float(value):
    return isinstance(value, (int, float))

emp = []



for k in keys_rows.keys():
    kl = []
    if k in df_keys.index:

        # Select the DataFrame corresponding to the index k
        selected_df = df_keys.loc[[k]]
        # try:
            
            # Check if each element in each column is int or float
        df_temp = (pd.DataFrame(df_keys.loc[k])).T
        result = df_temp.applymap(is_int_or_float)

        
        # # Check if any column contains int or float values
        column_contains_int_or_float = result.any()

        # df_temp = (pd.DataFrame(column_contains_int_or_float)).T
        for col in df_temp.columns:
            if column_contains_int_or_float[col]:
                kl.append(col)
        emp.append(kl)
        dict_another.append( df_keys.loc[[k], kl].dropna(axis=1))



for k in range(len(dict_another)):
    for j in range(len(dict_another[k].values[0])):
        
        if str(day_month[dict_another[k].columns].values[0][j]) in data.keys():
            data[(str(day_month[dict_another[k].columns].values[0][j]))].append([dict_another[k].index[0], dict_another[k].values[0][j]])
                
            
        if not str(day_month[dict_another[k].columns].values[0][j]) in data.keys():
            data[(str(day_month[dict_another[k].columns].values[0][j]))] = [[dict_another[k].index[0], dict_another[k].values[0][j]]]
        
        
# =============================================================================
# if you want to add weekends uncomment below (uncomment means delete the # below, not in this line)
# =============================================================================
# for l in weekende_date:
#         l = str(l)
#         if l not in data.keys():
#             data[l] = [['Weekend', 0.0]]
            

        
sorted_dict = {key: data[key] for key in sorted(data, key=lambda x: int(x))}

data = sorted_dict




# Iterate through the original dictionary
for day, projects in data.items():
    # Calculate the total hours for this day
    total_hours = sum(hours for _, hours in projects)

    # Initialize the starting hour
    if int(day) in day_of_second_change:
        current_hour = 15
        # Create a list to store the data for this day
        day_data = []

        # Iterate through the projects for this day
        for project_name, project_hours in projects:
            # Calculate the hours to allocate to this project
            allocated_hours = min(project_hours, total_hours)

            # Check if there are remaining hours to allocate
            if allocated_hours > 0:
                # Calculate the ending hour for this allocation
                end_hour = min(current_hour + allocated_hours, 23)

                # Add the allocation to the day_data list
                day_data.append(
                    {'Data': day, 'Projekt': project_name,
                                 'Godzina Od': f'{current_hour}', 
                                 'Godzina Do':f'{end_hour}'}
                                )

                # Update the current hour and remaining hours
                current_hour = end_hour
                total_hours -= allocated_hours

        # Create a DataFrame for this day and append it to the list of DataFrames
        if day_data:
            day_df = pd.DataFrame(day_data)
            dfs.append(day_df)
            
        
    else:
        current_hour = 7

        # Create a list to store the data for this day
        day_data = []
        
        # Iterate through the projects for this day
        for project_name, project_hours in projects:
            # Calculate the hours to allocate to this project
            allocated_hours = min(project_hours, total_hours)
    
            # Check if there are remaining hours to allocate
            if allocated_hours == 0:
                day_data.append(
                    {'Data': day, 'Projekt': project_name,
                                 'Godzina Od': '0', 
                                 'Godzina Do':'0'}
                                )
    
                
            
            if allocated_hours > 0:
                # Calculate the ending hour for this allocation
                end_hour = min(current_hour + allocated_hours, 15)
    
                # Add the allocation to the day_data list
                day_data.append(
                    {'Data': day, 'Projekt': project_name,
                                 'Godzina Od': f'{current_hour}', 
                                 'Godzina Do':f'{end_hour}'}
                                )
    
                # Update the current hour and remaining hours
                current_hour = end_hour
                total_hours -= allocated_hours
    
        # Create a DataFrame for this day and append it to the list of DataFrames
        if day_data:
            day_df = pd.DataFrame(day_data)
            dfs.append(day_df)

# Concatenate all DataFrames into a single result DataFrame
result_df = pd.concat(dfs, ignore_index=True)
result_copy = result_df.copy()


val_to_add = []
for _ in range(len(result_copy)):
    
    val = float(result_copy['Godzina Do'].iloc[_])-float(result_copy['Godzina Od'].iloc[_])
    val_to_add.append(val)


sum_dict = {}

for k in range(len(dict_another)):
    name_sum = 'Sum of hours - ' + dict_another[k].index[0]
    sum_dict[name_sum] = [sum(dict_another[k].values[0])]


df_sum = pd.DataFrame.from_dict(sum_dict)

df_hour_project_final = pd.DataFrame([hours_project], columns=names_of_projects)


time_modifier = TimeFormatModifier(column_name='Godzina Od')
time_modifier.modify_column(result_df)

time_modifier = TimeFormatModifier(column_name='Godzina Do')
time_modifier.modify_column(result_df)

dict_val = {'Liczba godzin':val_to_add}
df_val = pd.DataFrame.from_dict(dict_val)

result_df = pd.concat([result_df, df_val], axis=1)
col_order = ['Data', 'Godzina Od', 'Godzina Do', 'Liczba godzin', 'Projekt']
result_df = result_df[col_order]

nr_ewid_len = [nr_ewid] * len(result_copy)
nr_ewid_col = pd.DataFrame({'Numer ewidencyjny': nr_ewid_len})

proper_date = []
for k in result_copy['Data']:
    
    if int(k) < 10:
        k = f"{current_year}.{current_month}.0{k}"
        proper_date.append(k)
    else:
        k = f'{current_year}.{current_month}.{k}'
        proper_date.append(k)        
        

names_col = [user_input] * len(result_copy)
names_df = pd.DataFrame({'Imię i Nazwisko':names_col})
final_df = pd.concat([names_df,nr_ewid_col, result_df, df_sum, df_hour_project_final], 
                     axis=1)

final_df['Data'] = proper_date



# Print the result DataFrame
print(final_df)


final_df.to_excel(pt_to_save+f'/{user_input}.xlsx', index=False)


# Open the existing Excel file
file_path = file_n
# Create a new sheet
new_sheet_name = f'Podsumowanie godzin {current_month}.{current_year}'
ind_to_save = roboczo_godziny[roboczo_godziny.index.map(lambda x: isinstance(x, str))]



df_pods = pd.DataFrame({f'Podsumowanie godzin {current_month}.{current_year}':list(ind_to_save.index), 
                        'Wykazane':ind_to_save.sum(axis=1)})


book = load_workbook(file_path)

if len(book.sheetnames) == 1:
    # Create a new Excel workbook and add the DataFrame to a new sheet
    writer = pd.ExcelWriter(file_path, engine='openpyxl')
    writer.book = book
    df_pods.to_excel(writer, sheet_name=new_sheet_name, index=False)
    
    # Save the changes to the Excel file
    writer.save()
    writer.close()

else:
    print("""Nie udało się dodać zakładki podsumowanie, co oznacza, że w Twoim pliku jest już zakładka z podsumowaniem godzin lub znajdują się dodatkowe zakładki, których nie powinno być.""")


